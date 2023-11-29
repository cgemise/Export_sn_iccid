
from openpyxl import Workbook
from netmiko import ConnectHandler,NetmikoTimeoutException, NetmikoAuthenticationException
from openpyxl import Workbook
import threading
from concurrent.futures import ThreadPoolExecutor
import logging
from dotenv import load_dotenv
import os

load_dotenv('variable.env')  # Charge les variables d'environnement depuis .env

INPUT_FILE_PATH = os.getenv('INPUT_FILE_PATH')
EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH')
LOG_FILE_PATH = os.getenv('LOG_FILE_PATH')

# Configuration du logger
logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s')



def connect_and_execute(device_info, command):
    """
    Établit une connexion SSH à un appareil et exécute une commande donnée.
    Utilise le mot de passe secret pour accéder au mode enable si nécessaire.
    Retourne la sortie de la commande.
    """
    try:
        # Générer le mot de passe et le mot de passe secret
        password = generate_password(device_info['slug'])
        secret = generate_password_enable(device_info['slug'])

        # Paramètres de connexion
        device = {
            'device_type': 'cisco_ios',
            'ip': device_info['ip'],
            'username': 'waycom',
            'password': password,
            'secret': secret,
            'timeout': 60,
        }

        # Connexion à l'appareil
        with ConnectHandler(**device) as connection:
            # Passer en mode enable si nécessaire
            if connection.check_enable_mode():
                connection.enable()

            # Exécuter la commande
            output = connection.send_command(command)
            return output,None


    except NetmikoTimeoutException:
        error_msg = f"Timeout : Équipement non joignable {device_info['ip']}"
        logging.error(error_msg)
        return None, 'timeout'

    except NetmikoAuthenticationException:
        error_msg = f"Erreur d'authentification pour l'appareil {device_info['ip']}"
        logging.error(error_msg)
        return None, 'auth_error'

    except Exception as e:
        error_msg = f"Erreur de connexion pour l'appareil {device_info['ip']}: {e}"
        logging.error(error_msg)
        return None, 'other_error'

    return None, 'unknown_error'




def read_input_file(file_path):
    """
    Lit un fichier contenant les informations des appareils. Gère deux formats :
    1. IP avec masque, ref_client, slug, nd
    2. IP avec masque1, IP avec masque2, ref_client, slug, nd
    Renvoie les informations sous forme de liste de dictionnaires, en extrayant l'adresse IP sans le masque.
    Gère également les lignes mal formatées.
    """
    devices = []
    line_count = 0
    with open(file_path, 'r') as file:
        for line in file:
            line_count += 1
            try:
                parts = line.strip().split()
                # Détecter le format de la ligne et extraire l'IP avec masque
                if len(parts) == 5:  # Format avec deux IP
                    ip_with_mask, _, ref_client, slug, nd = parts
                elif len(parts) == 4:  # Format standard avec une IP
                    ip_with_mask, ref_client, slug, nd = parts
                else:
                    raise ValueError("Nombre incorrect de champs dans la ligne")

                ip = ip_with_mask.split('/')[0]  # Sépare l'IP du masque et prend l'IP
                devices.append({'ip': ip, 'ref_client': ref_client, 'slug': slug, 'nd': nd})
                logging.info(f"Ligne traitée avec succès : {line.strip()}")
            except ValueError as e:
                error_msg = f"Erreur de format de ligne : '{line.strip()}'. Erreur : {e}"
                print(error_msg)
                logging.error(error_msg)
    return devices, line_count



def generate_password_enable(slug):
    """
    Génère un mot de passe en suivant la nomenclature donnée.
    """
    return slug[0] + slug[-1] + 'dusuka!'

def generate_password(slug):
    """
    Génère un mot de passe en suivant la nomenclature donnée.
    """
    return 'dusuka' + slug[0] + slug[-1] + '!'


def handle_device(device_info):
    logging.debug(f"Traitement de l'appareil: {device_info['ip']}")
    """
    Gère la connexion à un appareil et récupère les données nécessaires (SN, ICCID).
    Retourne un dictionnaire contenant les informations collectées et les détails des erreurs, le cas échéant.
    """

    error = False
    error_msg = ""
    sn, iccid = None, None



    # Commandes pour récupérer le SN et identifier les appareils 4G
    sn_command = 'sh version | i Processor board ID'
    identify_4g_command = 'sh version | i bytes of memory'

    # Connexion à l'appareil et récupération du SN
    sn_output, error_type = connect_and_execute(device_info, sn_command)
    if sn_output and not error_type: 
        logging.debug(f"SN récupéré pour {device_info['ip']}: {sn_output}")
        try:
            sn = sn_output.split()[3]  # SN est à la 4ème position
            logging.info(f"SN récupéré pour l'appareil {device_info['ip']}: {sn}")
        except IndexError:
            error = True
            error_msg = f"Impossible de récupérer le SN pour l'appareil {device_info['ip']}"
            logging.error(error_msg)
            print(error_msg)
    else:
        logging.debug(f"Erreur détectée pour {device_info['ip']}, Type d'erreur: {error_type}")
        error = True
        if error_type == 'timeout':
            error_msg = f"Timeout pour l'appareil {device_info['ip']}"
        elif error_type == 'auth_error':
            error_msg = f"Erreur d'authentification pour l'appareil {device_info['ip']}"
        else:
            error_msg = f"Autre erreur pour l'appareil {device_info['ip']}"

        logging.error(error_msg)
        print(error_msg)

    if not error:

        # Identification du type d'appareil 4G et récupération de l'ICCID si nécessaire
        identify_4g_output, error_type_4g = connect_and_execute(device_info, identify_4g_command)
        #iccid = None
        if identify_4g_output and not error_type_4g:
            model = identify_4g_output.split()[1]  # Le modèle est à la 2ème position
            logging.info(f"Modèle d'appareil 4G identifié pour {device_info['ip']}: {model}")

            if model == 'C881G-4G-GA-K9':
                iccid_command = 'sh cellular 0 all | i ICCID'
                iccid_output, error_type_iccid = connect_and_execute(device_info, iccid_command)
                if iccid_output and not error_type_iccid:
                    try:
                        iccid = iccid_output.split()[6]  # ICCID est à la 7ème position
                        logging.info(f"ICCID récupéré pour l'appareil {device_info['ip']}: {iccid}")
                    except IndexError:
                        error = True
                        error_msg = f"Impossible de récupérer l'ICCID pour l'appareil {device_info['ip']}"
                        logging.error(error_msg)
                        print(error_msg)

                
            elif model == 'C1111-4PLTEEA':
                iccid_command = 'sh Cellular 0/2/0 all | i ICCID'
                iccid_output, error_type_iccid = connect_and_execute(device_info, iccid_command)
                if iccid_output and not error_type_iccid:
                    try:
                        iccid = iccid_output.split()[6]  # ICCID est à la 7ème position
                        logging.info(f"ICCID récupéré pour l'appareil {device_info['ip']}: {iccid}")
                    except IndexError:
                        error = True
                        error_msg = f"Impossible de récupérer l'ICCID pour l'appareil {device_info['ip']}"
                        logging.error(error_msg)
                        print(error_msg)
        pass
    logging.debug(f"Fin du traitement pour {device_info['ip']}, SN: {sn}, ICCID: {iccid}")
    # Retourne les informations récupérées
    return {'ip': device_info['ip'], 'ref_client': device_info['ref_client'], 
            'slug': device_info['slug'], 'nd': device_info['nd'], 
            'sn': sn, 'iccid': iccid, 'error': error, 'error_msg': error_msg}



def create_excel_file(devices_data,  total_lines):
    logging.debug("Création du fichier Excel")
    """
    Crée un fichier Excel avec trois feuilles distinctes basées sur les données fournies.
    """
    wb = Workbook()
    sn_sheet = wb.active
    sn_sheet.title = "Devices avec SN"
    g4_sheet = wb.create_sheet("Devices 4G")
    error_sheet = wb.create_sheet("Devices avec Erreurs")

    # Ajout des en-têtes pour chaque feuille
    sn_sheet.append(['ip', 'ref_client', 'slug', 'nd', 'sn'])
    g4_sheet.append(['ip', 'ref_client', 'slug', 'nd', 'sn', 'iccid'])
    error_sheet.append(['ip', 'ref_client', 'slug', 'nd', 'raison'])

    # Remplissage des feuilles avec les données
    for device in devices_data:
        logging.debug(f"Ajout de l'appareil au fichier Excel: {device}")
        if device['error']:
            error_sheet.append([device['ip'], device['ref_client'], device['slug'], device['nd'], device['error_msg']])
        elif device['iccid']:
            g4_sheet.append([device['ip'], device['ref_client'], device['slug'], device['nd'], device['sn'], device['iccid']])
        else:
            sn_sheet.append([device['ip'], device['ref_client'], device['slug'], device['nd'], device['sn']])

    logging.debug("Fichier Excel créé avec succès")

    # Compte rendu
    sn_count = len([d for d in devices_data if not d['error'] and not d['iccid']])
    g4_count = len([d for d in devices_data if d['iccid']])
    error_count = len([d for d in devices_data if d['error']])

    report_sheet = wb.create_sheet("Compte Rendu")
    report_sheet.append(["Total lignes dans le fichier d'entrée", total_lines])
    report_sheet.append(["Nombre de Devices avec SN", sn_count])
    report_sheet.append(["Nombre de Devices 4G", g4_count])
    report_sheet.append(["Nombre de Devices avec Erreurs", error_count])

    # Affichage des informations dans la console
    print(f"Total lignes dans le fichier d'entrée: {total_lines}")
    print(f"Nombre de Devices avec SN: {sn_count}")
    print(f"Nombre de Devices 4G: {g4_count}")
    print(f"Nombre de Devices avec Erreurs: {error_count}")

    wb.save(EXCEL_FILE_PATH)






def main():
    device_list, total_lines = read_input_file(INPUT_FILE_PATH)
    processed_devices = []  # Liste pour stocker les données de chaque appareil

    # Nombre maximum de threads simultanés
    max_threads = 75  # Vous pouvez ajuster ce nombre selon vos besoins et ressources

    with ThreadPoolExecutor(max_workers=max_threads) as executor:
        # Créer un futur pour chaque appareil dans la liste
        futures = [executor.submit(handle_device, device) for device in device_list]

        # Attendre que chaque futur soit complété et collecter les résultats
        for future in futures:
            processed_devices.append(future.result())

    # Appeler la fonction de création de fichier Excel avec les données collectées
    create_excel_file(processed_devices, total_lines)
    

if __name__ == "__main__":
    main()




